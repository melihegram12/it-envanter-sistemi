import os
import uuid
import hashlib
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from typing import List, Optional
from models import *

class ExcelManager:
    def __init__(self, file_path: str = "inventory_data.xlsx"):
        self.file_path = file_path
        self.ensure_file_exists()
    
    def ensure_file_exists(self):
        """Excel dosyası yoksa oluştur ve başlık satırlarını ekle"""
        if not os.path.exists(self.file_path):
            wb = Workbook()
            
            # Kullanıcılar
            ws_users = wb.active
            ws_users.title = "Kullanicilar"
            self._set_headers(ws_users, ["Username", "Şifre", "Ad Soyad", "Email", "Departman", "Rol", "Aktif", "Son Giriş"])
            
            # Malzemeler
            ws_materials = wb.create_sheet("Malzemeler")
            self._set_headers(ws_materials, ["Kod", "Ad", "Kategori", "Birim", "Mevcut Stok", "Min Seviye", "Max Seviye", "Konum", "Raf", "Barkod", "Birim Fiyat", "Son Güncelleme", "Son Sayım"])
            
            # Hareketler
            ws_movements = wb.create_sheet("Hareketler")
            self._set_headers(ws_movements, ["Tarih", "Malzeme Kodu", "İşlem Tipi", "Miktar", "Tedarikçi/Teslim Alan", "Açıklama", "Sipariş No", "Onaylayan"])
            
            # Tedarikçiler
            ws_suppliers = wb.create_sheet("Tedarikciler")
            self._set_headers(ws_suppliers, ["Kod", "Ad", "Yetkili Kişi", "Telefon", "Email", "Adres", "Kategori", "Puan", "Notlar", "Son Sipariş", "Toplam Sipariş", "Aktif"])
            
            # Siparişler
            ws_orders = wb.create_sheet("Siparisler")
            self._set_headers(ws_orders, ["Sipariş No", "Tarih", "Tedarikçi Kodu", "Tedarikçi Adı", "Durum", "Toplam Tutar", "Oluşturan", "Onaylayan", "Tahmini Teslim", "Teslim Tarihi", "Notlar", "Kalemler"])
            
            # Talepler
            ws_requests = wb.create_sheet("Talepler")
            self._set_headers(ws_requests, ["Talep No", "Tarih", "Malzeme Kodu", "Malzeme Adı", "Miktar", "Öncelik", "Talep Eden", "Departman", "Durum", "Onaylayan", "Onay Tarihi", "Red Nedeni", "Açıklama"])
            
            # Bütçe
            ws_budget = wb.create_sheet("Butce")
            self._set_headers(ws_budget, ["Yıl", "Kategori", "Aylık Limit", "Yıllık Limit", "Kullanılan", "Kalan"])
            
            # Bildirimler
            ws_notifications = wb.create_sheet("Bildirimler")
            self._set_headers(ws_notifications, ["ID", "Tarih", "Kullanıcı", "Tip", "Başlık", "Mesaj", "Link", "Okundu"])
            
            # Örnek verileri ekle
            self._add_sample_data(wb)
            
            wb.save(self.file_path)
    
    def _set_headers(self, ws, headers):
        """Başlık satırını formatla"""
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 18
    
    def _hash_password(self, password: str) -> str:
        return hashlib.sha256(password.encode()).hexdigest()
    
    def _generate_id(self, prefix: str = "") -> str:
        return f"{prefix}{datetime.now().strftime('%Y%m%d%H%M%S')}{str(uuid.uuid4())[:4].upper()}"
    
    def _add_sample_data(self, wb):
        """Örnek veriler ekle"""
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Kullanıcılar
        ws_users = wb["Kullanicilar"]
        sample_users = [
            ["admin", self._hash_password("admin123"), "Sistem Yöneticisi", "admin@sirket.com", "IT", "Admin", True, now],
            ["yonetici", self._hash_password("yonetici123"), "Ahmet Yılmaz", "ahmet@sirket.com", "İdari İşler", "Yönetici", True, now],
            ["kullanici", self._hash_password("kullanici123"), "Ayşe Demir", "ayse@sirket.com", "Muhasebe", "Kullanıcı", True, now],
        ]
        for row in sample_users:
            ws_users.append(row)
        
        # Malzemeler
        ws_materials = wb["Malzemeler"]
        sample_materials = [
            ["MAL001", "A4 Fotokopi Kağıdı", "Kırtasiye", "Paket", 50, 10, 200, "Depo A", "A-1", "8690000000001", 45.00, now, now],
            ["MAL002", "Tükenmez Kalem (Mavi)", "Kırtasiye", "Adet", 100, 20, 500, "Depo A", "A-2", "8690000000002", 5.50, now, now],
            ["MAL003", "Çok Amaçlı Temizleyici", "Temizlik", "Litre", 8, 5, 50, "Depo B", "B-1", "8690000000003", 35.00, now, now],
            ["MAL004", "Toner HP 26A", "Ofis Ekipmanı", "Adet", 3, 2, 10, "Depo C", "C-1", "8690000000004", 850.00, now, now],
            ["MAL005", "Kahve (Filtre)", "Mutfak", "Paket", 5, 3, 20, "Mutfak", "M-1", "8690000000005", 120.00, now, now],
            ["MAL006", "Plastik Bardak", "Mutfak", "Adet", 200, 50, 1000, "Mutfak", "M-2", "8690000000006", 0.50, now, now],
            ["MAL007", "Zımba Teli", "Kırtasiye", "Kutu", 15, 5, 50, "Depo A", "A-3", "8690000000007", 12.00, now, now],
            ["MAL008", "Klasör (Geniş)", "Kırtasiye", "Adet", 25, 10, 100, "Depo A", "A-4", "8690000000008", 28.00, now, now],
        ]
        for row in sample_materials:
            ws_materials.append(row)
        
        # Tedarikçiler
        ws_suppliers = wb["Tedarikciler"]
        sample_suppliers = [
            ["TED001", "ABC Kırtasiye A.Ş.", "Mehmet Kaya", "0212 555 1234", "satis@abckirtasiye.com", "İstanbul", "Kırtasiye", 4.5, "", now, 15, True],
            ["TED002", "Temizlik Market Ltd.", "Fatma Öz", "0212 555 5678", "info@temizlikmarket.com", "İstanbul", "Temizlik", 4.0, "", now, 8, True],
            ["TED003", "Ofis Teknik", "Ali Demir", "0216 444 3333", "destek@ofisteknik.com", "Ankara", "Ofis Ekipmanı", 4.8, "Hızlı teslimat", now, 12, True],
        ]
        for row in sample_suppliers:
            ws_suppliers.append(row)
        
        # Hareketler
        ws_movements = wb["Hareketler"]
        sample_movements = [
            [now, "MAL001", "Giriş", 100, "ABC Kırtasiye A.Ş.", "Aylık sipariş", "SIP001", "Ahmet Yılmaz"],
            [now, "MAL001", "Çıkış", 50, "Muhasebe Birimi", "Departman talebi", "", "Ayşe Demir"],
            [now, "MAL003", "Giriş", 10, "Temizlik Market Ltd.", "Haftalık sipariş", "SIP002", "Ahmet Yılmaz"],
        ]
        for row in sample_movements:
            ws_movements.append(row)
        
        # Talepler
        ws_requests = wb["Talepler"]
        sample_requests = [
            ["TLP001", now, "MAL001", "A4 Fotokopi Kağıdı", 20, "Normal", "Ayşe Demir", "Muhasebe", "Beklemede", "", "", "", "Aylık ihtiyaç"],
            ["TLP002", now, "MAL004", "Toner HP 26A", 2, "Yüksek", "Mehmet Can", "IT", "Beklemede", "", "", "", "Yazıcı toneri bitti"],
        ]
        for row in sample_requests:
            ws_requests.append(row)
        
        # Siparişler
        ws_orders = wb["Siparisler"]
        sample_orders = [
            ["SIP001", now, "TED001", "ABC Kırtasiye A.Ş.", "Teslim Edildi", 4500.00, "Ahmet Yılmaz", "Sistem Yöneticisi", now, now, "Aylık sipariş", "MAL001:100:45.00"],
            ["SIP002", now, "TED002", "Temizlik Market Ltd.", "Yolda", 350.00, "Ahmet Yılmaz", "", now, "", "Haftalık", "MAL003:10:35.00"],
        ]
        for row in sample_orders:
            ws_orders.append(row)
        
        # Bütçe
        ws_budget = wb["Butce"]
        current_year = datetime.now().year
        sample_budget = [
            [current_year, "Kırtasiye", 5000, 60000, 12500, 47500],
            [current_year, "Temizlik", 3000, 36000, 8400, 27600],
            [current_year, "Ofis Ekipmanı", 10000, 120000, 25500, 94500],
            [current_year, "Mutfak", 2000, 24000, 6000, 18000],
            [current_year, "Teknik", 5000, 60000, 15000, 45000],
        ]
        for row in sample_budget:
            ws_budget.append(row)
        
        # Bildirimler
        ws_notifications = wb["Bildirimler"]
        sample_notifications = [
            [self._generate_id("BLD"), now, "admin", "Kritik Stok", "Toner HP 26A stokta azaldı", "Mevcut: 3, Minimum: 2", "/materials", False],
            [self._generate_id("BLD"), now, "yonetici", "Talep Onayı", "Yeni talep onayınızı bekliyor", "TLP001 - A4 Fotokopi Kağıdı", "/requests", False],
        ]
        for row in sample_notifications:
            ws_notifications.append(row)

    # ==================== KULLANICI İŞLEMLERİ ====================
    
    def authenticate_user(self, username: str, password: str) -> Optional[User]:
        """Kullanıcı girişi"""
        wb = load_workbook(self.file_path)
        ws = wb["Kullanicilar"]
        hashed = self._hash_password(password)
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == username and row[1] == hashed and row[6]:
                # Son giriş güncelle
                ws.cell(row=row_idx, column=8, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
                wb.save(self.file_path)
                wb.close()
                return User(
                    username=row[0],
                    ad_soyad=row[2],
                    email=row[3],
                    departman=row[4],
                    rol=row[5],
                    aktif=row[6],
                    son_giris=datetime.now().strftime("%Y-%m-%d %H:%M")
                )
        wb.close()
        return None
    
    def get_all_users(self) -> List[User]:
        """Tüm kullanıcıları getir"""
        wb = load_workbook(self.file_path)
        ws = wb["Kullanicilar"]
        users = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                users.append(User(
                    username=row[0],
                    ad_soyad=row[2],
                    email=row[3],
                    departman=row[4],
                    rol=row[5],
                    aktif=row[6],
                    son_giris=row[7] or ""
                ))
        wb.close()
        return users
    
    def create_user(self, user: UserCreate) -> User:
        """Yeni kullanıcı oluştur"""
        wb = load_workbook(self.file_path)
        ws = wb["Kullanicilar"]
        
        ws.append([
            user.username,
            self._hash_password(user.password),
            user.ad_soyad,
            user.email,
            user.departman,
            user.rol.value,
            user.aktif,
            ""
        ])
        wb.save(self.file_path)
        wb.close()
        return User(**user.dict(exclude={'password'}))

    # ==================== MALZEME İŞLEMLERİ ====================
    
    def get_all_materials(self) -> List[Material]:
        """Tüm malzemeleri getir"""
        wb = load_workbook(self.file_path)
        ws = wb["Malzemeler"]
        materials = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                mevcut = int(row[4]) if row[4] else 0
                min_s = int(row[5]) if row[5] else 0
                max_s = int(row[6]) if row[6] else 100
                
                if mevcut <= min_s:
                    durum = "Kritik"
                elif mevcut >= max_s:
                    durum = "Fazla"
                else:
                    durum = "Normal"
                
                materials.append(Material(
                    kod=row[0],
                    ad=row[1],
                    kategori=row[2],
                    birim=row[3],
                    mevcut_stok=mevcut,
                    min_seviye=min_s,
                    max_seviye=max_s,
                    konum=row[7] or "",
                    raf=row[8] or "",
                    barkod=row[9] or "",
                    birim_fiyat=row[10] or 0.0,
                    son_guncelleme=row[11] or "",
                    son_sayim=row[12] or "",
                    durum=durum
                ))
        wb.close()
        return materials
    
    def get_material_by_code(self, kod: str) -> Optional[Material]:
        """Kod ile malzeme bul"""
        materials = self.get_all_materials()
        for m in materials:
            if m.kod == kod:
                return m
        return None
    
    def get_material_by_barcode(self, barcode: str) -> Optional[Material]:
        """Barkod/QR kod ile malzeme bul"""
        materials = self.get_all_materials()
        for m in materials:
            if m.barkod and m.barkod.lower() == barcode.lower():
                return m
        return None
    
    def create_material(self, material: MaterialCreate) -> Material:
        """Yeni malzeme ekle"""
        wb = load_workbook(self.file_path)
        ws = wb["Malzemeler"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        ws.append([
            material.kod,
            material.ad,
            material.kategori.value,
            material.birim.value,
            material.mevcut_stok,
            material.min_seviye,
            material.max_seviye,
            material.konum,
            material.raf,
            material.barkod,
            material.birim_fiyat,
            now,
            now
        ])
        wb.save(self.file_path)
        wb.close()
        return Material(**material.dict(), son_guncelleme=now, son_sayim=now, durum="Normal")
    
    def update_material(self, kod: str, material: MaterialCreate) -> Optional[Material]:
        """Malzeme güncelle"""
        wb = load_workbook(self.file_path)
        ws = wb["Malzemeler"]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == kod:
                now = datetime.now().strftime("%Y-%m-%d %H:%M")
                ws.cell(row=row_idx, column=1, value=material.kod)
                ws.cell(row=row_idx, column=2, value=material.ad)
                ws.cell(row=row_idx, column=3, value=material.kategori.value)
                ws.cell(row=row_idx, column=4, value=material.birim.value)
                ws.cell(row=row_idx, column=5, value=material.mevcut_stok)
                ws.cell(row=row_idx, column=6, value=material.min_seviye)
                ws.cell(row=row_idx, column=7, value=material.max_seviye)
                ws.cell(row=row_idx, column=8, value=material.konum)
                ws.cell(row=row_idx, column=9, value=material.raf)
                ws.cell(row=row_idx, column=10, value=material.barkod)
                ws.cell(row=row_idx, column=11, value=material.birim_fiyat)
                ws.cell(row=row_idx, column=12, value=now)
                wb.save(self.file_path)
                wb.close()
                return Material(**material.dict(), son_guncelleme=now, durum="Normal")
        wb.close()
        return None
    
    def delete_material(self, kod: str) -> bool:
        """Malzeme sil"""
        wb = load_workbook(self.file_path)
        ws = wb["Malzemeler"]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == kod:
                ws.delete_rows(row_idx)
                wb.save(self.file_path)
                wb.close()
                return True
        wb.close()
        return False

    # ==================== STOK HAREKETLERİ ====================
    
    def get_all_movements(self) -> List[StockMovement]:
        """Tüm hareketleri getir"""
        wb = load_workbook(self.file_path)
        ws = wb["Hareketler"]
        movements = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                movements.append(StockMovement(
                    tarih=str(row[0]),
                    malzeme_kodu=row[1],
                    islem_tipi=row[2],
                    miktar=row[3] or 0,
                    tedarikci_teslim_alan=row[4] or "",
                    aciklama=row[5] or "",
                    siparis_no=row[6] or "",
                    onaylayan=row[7] or ""
                ))
        wb.close()
        return list(reversed(movements))
    
    def create_movement(self, movement: StockMovementCreate) -> StockMovement:
        """Yeni hareket ekle ve stok güncelle"""
        wb = load_workbook(self.file_path)
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Hareketi ekle
        ws_movements = wb["Hareketler"]
        ws_movements.append([
            now,
            movement.malzeme_kodu,
            movement.islem_tipi.value,
            movement.miktar,
            movement.tedarikci_teslim_alan,
            movement.aciklama,
            movement.siparis_no,
            movement.onaylayan
        ])
        
        # Stoku güncelle
        ws_materials = wb["Malzemeler"]
        for row_idx, row in enumerate(ws_materials.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == movement.malzeme_kodu:
                current_stock = row[4] or 0
                if movement.islem_tipi == MovementType.GIRIS:
                    new_stock = current_stock + movement.miktar
                else:
                    new_stock = max(0, current_stock - movement.miktar)
                
                ws_materials.cell(row=row_idx, column=5, value=new_stock)
                ws_materials.cell(row=row_idx, column=12, value=now)
                
                # Kritik stok kontrolü
                min_level = row[5] or 0
                if new_stock <= min_level:
                    self._create_notification_internal(wb, NotificationCreate(
                        kullanici="admin",
                        tip=NotificationType.CRITICAL_STOCK,
                        baslik=f"{row[1]} kritik stok seviyesinde",
                        mesaj=f"Mevcut: {new_stock}, Minimum: {min_level}",
                        link="/materials"
                    ))
                break
        
        wb.save(self.file_path)
        wb.close()
        return StockMovement(**movement.dict(), tarih=now)

    # ==================== TEDARİKÇİ İŞLEMLERİ ====================
    
    def get_all_suppliers(self) -> List[Supplier]:
        """Tüm tedarikçileri getir"""
        wb = load_workbook(self.file_path)
        ws = wb["Tedarikciler"]
        suppliers = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                suppliers.append(Supplier(
                    kod=row[0],
                    ad=row[1],
                    yetkili_kisi=row[2] or "",
                    telefon=row[3] or "",
                    email=row[4] or "",
                    adres=row[5] or "",
                    kategori=row[6] or "",
                    puan=row[7] or 5.0,
                    notlar=row[8] or "",
                    son_siparis=row[9] or "",
                    toplam_siparis=row[10] or 0,
                    aktif=row[11] if row[11] is not None else True
                ))
        wb.close()
        return suppliers
    
    def get_supplier_by_code(self, kod: str) -> Optional[Supplier]:
        """Kod ile tedarikçi bul"""
        suppliers = self.get_all_suppliers()
        for s in suppliers:
            if s.kod == kod:
                return s
        return None
    
    def create_supplier(self, supplier: SupplierCreate) -> Supplier:
        """Yeni tedarikçi ekle"""
        wb = load_workbook(self.file_path)
        ws = wb["Tedarikciler"]
        
        ws.append([
            supplier.kod,
            supplier.ad,
            supplier.yetkili_kisi,
            supplier.telefon,
            supplier.email,
            supplier.adres,
            supplier.kategori,
            supplier.puan,
            supplier.notlar,
            "",
            0,
            True
        ])
        wb.save(self.file_path)
        wb.close()
        return Supplier(**supplier.dict(), son_siparis="", toplam_siparis=0, aktif=True)
    
    def update_supplier(self, kod: str, supplier: SupplierCreate) -> Optional[Supplier]:
        """Tedarikçi güncelle"""
        wb = load_workbook(self.file_path)
        ws = wb["Tedarikciler"]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == kod:
                ws.cell(row=row_idx, column=1, value=supplier.kod)
                ws.cell(row=row_idx, column=2, value=supplier.ad)
                ws.cell(row=row_idx, column=3, value=supplier.yetkili_kisi)
                ws.cell(row=row_idx, column=4, value=supplier.telefon)
                ws.cell(row=row_idx, column=5, value=supplier.email)
                ws.cell(row=row_idx, column=6, value=supplier.adres)
                ws.cell(row=row_idx, column=7, value=supplier.kategori)
                ws.cell(row=row_idx, column=8, value=supplier.puan)
                ws.cell(row=row_idx, column=9, value=supplier.notlar)
                wb.save(self.file_path)
                wb.close()
                return Supplier(**supplier.dict(), son_siparis=row[9] or "", toplam_siparis=row[10] or 0, aktif=True)
        wb.close()
        return None
    
    def delete_supplier(self, kod: str) -> bool:
        """Tedarikçi sil"""
        wb = load_workbook(self.file_path)
        ws = wb["Tedarikciler"]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == kod:
                ws.delete_rows(row_idx)
                wb.save(self.file_path)
                wb.close()
                return True
        wb.close()
        return False

    # ==================== SİPARİŞ İŞLEMLERİ ====================
    
    def get_all_orders(self) -> List[Order]:
        """Tüm siparişleri getir"""
        wb = load_workbook(self.file_path)
        ws = wb["Siparisler"]
        orders = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                # Kalemleri parse et
                kalemler = []
                if row[11]:
                    for item in str(row[11]).split(";"):
                        parts = item.split(":")
                        if len(parts) == 3:
                            kalemler.append(OrderItem(
                                malzeme_kodu=parts[0],
                                malzeme_adi=parts[0],
                                miktar=int(parts[1]),
                                birim_fiyat=float(parts[2]),
                                toplam=int(parts[1]) * float(parts[2])
                            ))
                
                orders.append(Order(
                    siparis_no=row[0],
                    tarih=str(row[1]),
                    tedarikci_kodu=row[2] or "",
                    tedarikci_adi=row[3] or "",
                    durum=row[4] or "Onay Bekliyor",
                    toplam_tutar=row[5] or 0.0,
                    olusturan=row[6] or "",
                    onaylayan=row[7] or "",
                    tahmini_teslim=row[8] or "",
                    teslim_tarihi=row[9] or "",
                    notlar=row[10] or "",
                    kalemler=kalemler
                ))
        wb.close()
        return list(reversed(orders))
    
    def create_order(self, order: OrderCreate) -> Order:
        """Yeni sipariş oluştur"""
        wb = load_workbook(self.file_path)
        ws = wb["Siparisler"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        siparis_no = self._generate_id("SIP")
        
        # Kalemleri string'e çevir
        kalem_str = ";".join([f"{k.malzeme_kodu}:{k.miktar}:{k.birim_fiyat}" for k in order.kalemler])
        toplam = sum(k.miktar * k.birim_fiyat for k in order.kalemler)
        
        ws.append([
            siparis_no,
            now,
            order.tedarikci_kodu,
            order.tedarikci_adi,
            "Onay Bekliyor",
            toplam,
            order.olusturan,
            "",
            "",
            "",
            order.notlar,
            kalem_str
        ])
        wb.save(self.file_path)
        wb.close()
        
        return Order(
            siparis_no=siparis_no,
            tarih=now,
            tedarikci_kodu=order.tedarikci_kodu,
            tedarikci_adi=order.tedarikci_adi,
            kalemler=order.kalemler,
            notlar=order.notlar,
            toplam_tutar=toplam,
            olusturan=order.olusturan,
            durum=OrderStatus.PENDING
        )
    
    def update_order_status(self, siparis_no: str, durum: OrderStatus, onaylayan: str = "") -> Optional[Order]:
        """Sipariş durumu güncelle"""
        wb = load_workbook(self.file_path)
        ws = wb["Siparisler"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == siparis_no:
                ws.cell(row=row_idx, column=5, value=durum.value)
                if onaylayan:
                    ws.cell(row=row_idx, column=8, value=onaylayan)
                if durum == OrderStatus.DELIVERED:
                    ws.cell(row=row_idx, column=10, value=now)
                wb.save(self.file_path)
                wb.close()
                return self.get_order_by_no(siparis_no)
        wb.close()
        return None
    
    def get_order_by_no(self, siparis_no: str) -> Optional[Order]:
        """Sipariş no ile sipariş bul"""
        orders = self.get_all_orders()
        for o in orders:
            if o.siparis_no == siparis_no:
                return o
        return None

    # ==================== TALEP İŞLEMLERİ ====================
    
    def get_all_requests(self) -> List[Request]:
        """Tüm talepleri getir"""
        wb = load_workbook(self.file_path)
        ws = wb["Talepler"]
        requests = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                requests.append(Request(
                    talep_no=row[0],
                    tarih=str(row[1]),
                    malzeme_kodu=row[2] or "",
                    malzeme_adi=row[3] or "",
                    miktar=row[4] or 0,
                    oncelik=row[5] or "Normal",
                    talep_eden=row[6] or "",
                    departman=row[7] or "",
                    durum=row[8] or "Beklemede",
                    onaylayan=row[9] or "",
                    onay_tarihi=row[10] or "",
                    red_nedeni=row[11] or "",
                    aciklama=row[12] or ""
                ))
        wb.close()
        return list(reversed(requests))
    
    def create_request(self, request: RequestCreate) -> Request:
        """Yeni talep oluştur"""
        wb = load_workbook(self.file_path)
        ws = wb["Talepler"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        talep_no = self._generate_id("TLP")
        
        ws.append([
            talep_no,
            now,
            request.malzeme_kodu,
            request.malzeme_adi,
            request.miktar,
            request.oncelik.value,
            request.talep_eden,
            request.departman,
            "Beklemede",
            "",
            "",
            "",
            request.aciklama
        ])
        
        # Bildirim oluştur
        self._create_notification_internal(wb, NotificationCreate(
            kullanici="yonetici",
            tip=NotificationType.SYSTEM,
            baslik="Yeni talep oluşturuldu",
            mesaj=f"{talep_no} - {request.malzeme_adi} ({request.miktar} adet)",
            link="/requests"
        ))
        
        wb.save(self.file_path)
        wb.close()
        
        return Request(
            talep_no=talep_no,
            tarih=now,
            malzeme_kodu=request.malzeme_kodu,
            malzeme_adi=request.malzeme_adi,
            miktar=request.miktar,
            oncelik=request.oncelik,
            talep_eden=request.talep_eden,
            departman=request.departman,
            durum=RequestStatus.PENDING,
            aciklama=request.aciklama
        )
    
    def update_request_status(self, talep_no: str, durum: RequestStatus, onaylayan: str = "", red_nedeni: str = "") -> Optional[Request]:
        """Talep durumu güncelle"""
        wb = load_workbook(self.file_path)
        ws = wb["Talepler"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == talep_no:
                ws.cell(row=row_idx, column=9, value=durum.value)
                ws.cell(row=row_idx, column=10, value=onaylayan)
                ws.cell(row=row_idx, column=11, value=now)
                if red_nedeni:
                    ws.cell(row=row_idx, column=12, value=red_nedeni)
                
                # Bildirim oluştur
                talep_eden = row[6]
                if durum == RequestStatus.APPROVED:
                    self._create_notification_internal(wb, NotificationCreate(
                        kullanici=talep_eden,
                        tip=NotificationType.REQUEST_APPROVED,
                        baslik="Talebiniz onaylandı",
                        mesaj=f"{talep_no} - {row[3]}",
                        link="/requests"
                    ))
                elif durum == RequestStatus.REJECTED:
                    self._create_notification_internal(wb, NotificationCreate(
                        kullanici=talep_eden,
                        tip=NotificationType.REQUEST_REJECTED,
                        baslik="Talebiniz reddedildi",
                        mesaj=f"{talep_no} - Neden: {red_nedeni}",
                        link="/requests"
                    ))
                
                wb.save(self.file_path)
                wb.close()
                return self.get_request_by_no(talep_no)
        wb.close()
        return None
    
    def get_request_by_no(self, talep_no: str) -> Optional[Request]:
        """Talep no ile talep bul"""
        requests = self.get_all_requests()
        for r in requests:
            if r.talep_no == talep_no:
                return r
        return None
    
    def get_pending_requests(self) -> List[Request]:
        """Bekleyen talepleri getir"""
        return [r for r in self.get_all_requests() if r.durum == "Beklemede"]

    # ==================== BÜTÇE İŞLEMLERİ ====================
    
    def get_budget_summary(self, yil: int = None) -> BudgetSummary:
        """Bütçe özeti"""
        if yil is None:
            yil = datetime.now().year
            
        wb = load_workbook(self.file_path)
        ws = wb["Butce"]
        budgets = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == yil:
                budgets.append(Budget(
                    yil=row[0],
                    kategori=row[1],
                    aylik_limit=row[2] or 0,
                    yillik_limit=row[3] or 0,
                    kullanilan=row[4] or 0,
                    kalan=row[5] or 0
                ))
        wb.close()
        
        toplam = sum(b.yillik_limit for b in budgets)
        kullanilan = sum(b.kullanilan for b in budgets)
        
        return BudgetSummary(
            yil=yil,
            toplam_butce=toplam,
            kullanilan=kullanilan,
            kalan=toplam - kullanilan,
            oran=(kullanilan / toplam * 100) if toplam > 0 else 0,
            kategoriler=budgets
        )
    
    def update_budget(self, yil: int, kategori: str, harcama: float):
        """Bütçe kullanımı güncelle"""
        wb = load_workbook(self.file_path)
        ws = wb["Butce"]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == yil and row[1] == kategori:
                kullanilan = (row[4] or 0) + harcama
                kalan = (row[3] or 0) - kullanilan
                ws.cell(row=row_idx, column=5, value=kullanilan)
                ws.cell(row=row_idx, column=6, value=kalan)
                
                # Limit aşımı kontrolü
                if kalan < 0:
                    self._create_notification_internal(wb, NotificationCreate(
                        kullanici="admin",
                        tip=NotificationType.BUDGET_WARNING,
                        baslik=f"{kategori} bütçe limiti aşıldı!",
                        mesaj=f"Aşım miktarı: {abs(kalan):.2f} TL",
                        link="/budget"
                    ))
                break
        
        wb.save(self.file_path)
        wb.close()

    # ==================== BİLDİRİM İŞLEMLERİ ====================
    
    def _create_notification_internal(self, wb, notification: NotificationCreate):
        """İç kullanım için bildirim oluştur (wb açık olmalı)"""
        ws = wb["Bildirimler"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        ws.append([
            self._generate_id("BLD"),
            now,
            notification.kullanici,
            notification.tip.value,
            notification.baslik,
            notification.mesaj,
            notification.link,
            False
        ])
    
    def create_notification(self, notification: NotificationCreate) -> Notification:
        """Bildirim oluştur"""
        wb = load_workbook(self.file_path)
        ws = wb["Bildirimler"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        notif_id = self._generate_id("BLD")
        
        ws.append([
            notif_id,
            now,
            notification.kullanici,
            notification.tip.value,
            notification.baslik,
            notification.mesaj,
            notification.link,
            False
        ])
        wb.save(self.file_path)
        wb.close()
        
        return Notification(
            id=notif_id,
            tarih=now,
            kullanici=notification.kullanici,
            tip=notification.tip,
            baslik=notification.baslik,
            mesaj=notification.mesaj,
            link=notification.link,
            okundu=False
        )
    
    def get_user_notifications(self, username: str) -> List[Notification]:
        """Kullanıcı bildirimlerini getir"""
        wb = load_workbook(self.file_path)
        ws = wb["Bildirimler"]
        notifications = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == username or row[2] == "all":
                notifications.append(Notification(
                    id=row[0],
                    tarih=str(row[1]),
                    kullanici=row[2],
                    tip=row[3],
                    baslik=row[4],
                    mesaj=row[5] or "",
                    link=row[6] or "",
                    okundu=row[7] or False
                ))
        wb.close()
        return list(reversed(notifications))
    
    def mark_notification_read(self, notif_id: str) -> bool:
        """Bildirimi okundu olarak işaretle"""
        wb = load_workbook(self.file_path)
        ws = wb["Bildirimler"]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == notif_id:
                ws.cell(row=row_idx, column=8, value=True)
                wb.save(self.file_path)
                wb.close()
                return True
        wb.close()
        return False

    # ==================== RAPORLAMA ====================
    
    def get_critical_stock_materials(self) -> List[Material]:
        """Kritik stok seviyesindeki malzemeler"""
        materials = self.get_all_materials()
        return [m for m in materials if m.mevcut_stok <= m.min_seviye]
    
    def get_category_distribution(self) -> dict:
        """Kategori bazlı dağılım"""
        materials = self.get_all_materials()
        distribution = {}
        for m in materials:
            cat = m.kategori
            if cat not in distribution:
                distribution[cat] = 0
            distribution[cat] += 1
        return distribution
    
    def get_total_stock_value(self) -> float:
        """Toplam stok değeri"""
        materials = self.get_all_materials()
        return sum(m.mevcut_stok * m.birim_fiyat for m in materials)
    
    def get_dashboard_stats(self, username: str = None) -> dict:
        """Dashboard istatistikleri"""
        materials = self.get_all_materials()
        movements = self.get_all_movements()
        requests = self.get_pending_requests()
        orders = [o for o in self.get_all_orders() if o.durum in ["Onay Bekliyor", "Yolda"]]
        budget = self.get_budget_summary()
        
        return {
            "toplam_malzeme": len(materials),
            "toplam_stok_degeri": self.get_total_stock_value(),
            "kritik_stok_sayisi": len(self.get_critical_stock_materials()),
            "bekleyen_talep_sayisi": len(requests),
            "bekleyen_siparis_sayisi": len(orders),
            "kategori_dagilimi": self.get_category_distribution(),
            "aylik_harcama": budget.kullanilan / 12 if budget.kullanilan else 0,
            "butce_durumu": {
                "toplam": budget.toplam_butce,
                "kullanilan": budget.kullanilan,
                "kalan": budget.kalan,
                "oran": budget.oran
            }
        }
    
    def get_department_consumption(self, departman: str = None) -> List[dict]:
        """Departman tüketim raporu"""
        movements = [m for m in self.get_all_movements() if m.islem_tipi == "Çıkış"]
        
        dept_data = {}
        for m in movements:
            dept = m.tedarikci_teslim_alan or "Bilinmiyor"
            if departman and dept != departman:
                continue
            if dept not in dept_data:
                dept_data[dept] = {"miktar": 0, "kalem": set()}
            dept_data[dept]["miktar"] += m.miktar
            dept_data[dept]["kalem"].add(m.malzeme_kodu)
        
        return [
            {"departman": k, "toplam_miktar": v["miktar"], "kalem_sayisi": len(v["kalem"])}
            for k, v in dept_data.items()
        ]

    # ==================== AUDIT LOG ====================
    
    def _ensure_audit_sheet(self, wb):
        """Audit log sayfasını kontrol et/oluştur"""
        if "AuditLog" not in wb.sheetnames:
            ws = wb.create_sheet("AuditLog")
            self._set_headers(ws, ["ID", "Tarih", "Kullanıcı", "İşlem", "Modül", "Kayıt ID", "Eski Değer", "Yeni Değer", "IP", "Detay"])
        return wb["AuditLog"]
    
    def create_audit_log(self, kullanici: str, islem: str, modul: str, kayit_id: str = "", eski: str = "", yeni: str = "", detay: str = ""):
        """Audit log oluştur"""
        wb = load_workbook(self.file_path)
        ws = self._ensure_audit_sheet(wb)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        ws.append([
            self._generate_id("AUD"),
            now,
            kullanici,
            islem,
            modul,
            kayit_id,
            eski[:200] if eski else "",
            yeni[:200] if yeni else "",
            "",
            detay
        ])
        wb.save(self.file_path)
        wb.close()
    
    def get_audit_logs(self) -> List[dict]:
        """Audit logları getir"""
        wb = load_workbook(self.file_path)
        if "AuditLog" not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb["AuditLog"]
        logs = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                logs.append({
                    "id": row[0],
                    "tarih": str(row[1]) if row[1] else "",
                    "kullanici": row[2] or "",
                    "islem": row[3] or "",
                    "modul": row[4] or "",
                    "kayit_id": row[5] or "",
                    "eski_deger": row[6] or "",
                    "yeni_deger": row[7] or "",
                    "ip_adresi": row[8] or "",
                    "detay": row[9] or ""
                })
        wb.close()
        return list(reversed(logs))

    # ==================== LOKASYONLAR ====================
    
    def _ensure_location_sheet(self, wb):
        """Lokasyon sayfasını kontrol et/oluştur"""
        if "Lokasyonlar" not in wb.sheetnames:
            ws = wb.create_sheet("Lokasyonlar")
            self._set_headers(ws, ["Kod", "Ad", "Adres", "Sorumlu", "Telefon", "Aktif"])
            # Örnek lokasyonlar
            ws.append(["LOK001", "Ana Depo", "İstanbul Merkez", "Ahmet Yılmaz", "0212 555 1234", True])
            ws.append(["LOK002", "Şube Depo", "Ankara", "Mehmet Kaya", "0312 444 5678", True])
        return wb["Lokasyonlar"]
    
    def get_all_locations(self) -> List[dict]:
        """Tüm lokasyonları getir"""
        wb = load_workbook(self.file_path)
        ws = self._ensure_location_sheet(wb)
        locations = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                locations.append({
                    "kod": row[0],
                    "ad": row[1] or "",
                    "adres": row[2] or "",
                    "sorumlu": row[3] or "",
                    "telefon": row[4] or "",
                    "aktif": row[5] if row[5] is not None else True
                })
        wb.save(self.file_path)
        wb.close()
        return locations
    
    def create_location(self, location) -> dict:
        """Yeni lokasyon ekle"""
        wb = load_workbook(self.file_path)
        ws = self._ensure_location_sheet(wb)
        
        ws.append([
            location.kod,
            location.ad,
            location.adres,
            location.sorumlu,
            location.telefon,
            location.aktif
        ])
        wb.save(self.file_path)
        wb.close()
        return {"kod": location.kod, "ad": location.ad}
    
    def delete_location(self, kod: str) -> bool:
        """Lokasyon sil"""
        wb = load_workbook(self.file_path)
        if "Lokasyonlar" not in wb.sheetnames:
            wb.close()
            return False
        
        ws = wb["Lokasyonlar"]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == kod:
                ws.delete_rows(row_idx)
                wb.save(self.file_path)
                wb.close()
                return True
        wb.close()
        return False

    # ==================== STOK SAYIM ====================
    
    def _ensure_count_sheet(self, wb):
        """Sayım sayfasını kontrol et/oluştur"""
        if "Sayimlar" not in wb.sheetnames:
            ws = wb.create_sheet("Sayimlar")
            self._set_headers(ws, ["Sayım No", "Tarih", "Lokasyon", "Kategori", "Durum", "Oluşturan", "Tamamlayan", "Tamamlanma", "Açıklama"])
        return wb["Sayimlar"]
    
    def get_all_stock_counts(self) -> List[dict]:
        """Tüm sayımları getir"""
        wb = load_workbook(self.file_path)
        ws = self._ensure_count_sheet(wb)
        counts = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                counts.append({
                    "sayim_no": row[0],
                    "tarih": str(row[1]) if row[1] else "",
                    "lokasyon": row[2] or "",
                    "kategori": row[3] or "",
                    "durum": row[4] or "Planlandı",
                    "olusturan": row[5] or "",
                    "tamamlayan": row[6] or "",
                    "tamamlanma_tarihi": str(row[7]) if row[7] else "",
                    "aciklama": row[8] or ""
                })
        wb.save(self.file_path)
        wb.close()
        return list(reversed(counts))
    
    def create_stock_count(self, count) -> dict:
        """Yeni sayım planla"""
        wb = load_workbook(self.file_path)
        ws = self._ensure_count_sheet(wb)
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        sayim_no = self._generate_id("SAY")
        
        ws.append([
            sayim_no,
            now,
            count.lokasyon,
            count.kategori,
            "Planlandı",
            count.olusturan,
            "",
            "",
            count.aciklama
        ])
        wb.save(self.file_path)
        wb.close()
        return {"sayim_no": sayim_no, "tarih": now, "durum": "Planlandı"}
    
    def complete_stock_count(self, sayim_no: str, tamamlayan: str) -> dict:
        """Sayımı tamamla"""
        wb = load_workbook(self.file_path)
        if "Sayimlar" not in wb.sheetnames:
            wb.close()
            return None
        
        ws = wb["Sayimlar"]
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == sayim_no:
                ws.cell(row=row_idx, column=5, value="Tamamlandı")
                ws.cell(row=row_idx, column=7, value=tamamlayan)
                ws.cell(row=row_idx, column=8, value=now)
                wb.save(self.file_path)
                wb.close()
                return {"sayim_no": sayim_no, "durum": "Tamamlandı"}
        wb.close()
        return None

    # ==================== ANALİTİK ====================
    
    # ==================== EXCEL İMPORT ====================
    
    def import_materials_from_excel(self, import_file_path: str) -> dict:
        """Kullanıcının Excel dosyasından malzeme içe aktar
        
        Beklenen kolonlar: Sayı, Ürün Adı, Adet, Durum
        Adet kolonu: "11 KUTU", "5 KUTU", "2 BİDON", "37" gibi değerler içerebilir
        """
        import re
        
        try:
            # Yüklenen Excel'i aç
            import_wb = load_workbook(import_file_path)
            import_ws = import_wb.active
            
            # Başlık satırını bul ve kolon indexlerini belirle
            headers = {}
            header_row = []
            
            for col_idx, cell in enumerate(import_ws[1], 1):
                cell_value = cell.value
                if cell_value:
                    header_lower = str(cell_value).lower().strip()
                    header_row.append(str(cell_value))
                    
                    # Sayı/Sıra/No kolonu
                    if 'sayı' in header_lower or 'sıra' in header_lower or 'no' in header_lower or header_lower == 's':
                        headers['sayi'] = col_idx
                    # Ürün Adı kolonu - daha fazla varyasyon
                    elif any(k in header_lower for k in ['ürün', 'urun', 'ad', 'malzeme', 'isim', 'adi', 'adı', 'product', 'name', 'ürün adı', 'urun adi']):
                        headers['urun_adi'] = col_idx
                    # Adet kolonu
                    elif any(k in header_lower for k in ['adet', 'miktar', 'stok', 'quantity', 'qty', 'amount']):
                        headers['adet'] = col_idx
                    # Durum kolonu
                    elif any(k in header_lower for k in ['durum', 'status', 'state']):
                        headers['durum'] = col_idx
                else:
                    header_row.append('')
            
            # Eğer urun_adi bulunamadıysa, 2. kolonu varsayılan olarak kullan
            if 'urun_adi' not in headers:
                # Kolon sayısı 2 veya daha fazlaysa, 2. kolonu ürün adı olarak kabul et
                if len(header_row) >= 2:
                    headers['urun_adi'] = 2
                else:
                    import_wb.close()
                    return {
                        "success": False, 
                        "error": f"Ürün Adı kolonu bulunamadı. Bulunan kolonlar: {', '.join(header_row) if header_row else 'Boş'}",
                        "imported": 0, 
                        "skipped": 0
                    }
            
            # Ana Excel dosyasını aç
            wb = load_workbook(self.file_path)
            ws = wb["Malzemeler"]
            now = datetime.now().strftime("%Y-%m-%d %H:%M")
            
            # Mevcut malzeme kodlarını al
            existing_codes = set()
            existing_names = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    existing_codes.add(row[0])
                if row[1]:
                    existing_names.add(row[1].lower().strip())
            
            imported_count = 0
            skipped_count = 0
            skipped_items = []
            
            # Verileri oku ve içe aktar
            for row_idx, row in enumerate(import_ws.iter_rows(min_row=2, values_only=True), start=2):
                # Ürün adını al
                urun_adi_col = headers.get('urun_adi', 2)
                urun_adi = row[urun_adi_col - 1] if len(row) >= urun_adi_col else None
                
                if not urun_adi or str(urun_adi).strip() == '':
                    continue
                
                urun_adi = str(urun_adi).strip()
                
                # Zaten var mı kontrol et
                if urun_adi.lower() in existing_names:
                    skipped_count += 1
                    skipped_items.append(urun_adi)
                    continue
                
                # Adet ve birim bilgisini parse et
                adet_col = headers.get('adet', 3)
                adet_raw = row[adet_col - 1] if len(row) >= adet_col else 0
                
                adet = 0
                birim = "Adet"
                
                if adet_raw:
                    adet_str = str(adet_raw).strip().upper()
                    # "11 KUTU", "5 KUTU", "2 BİDON", "8 BOY", "37" gibi formatları parse et
                    match = re.match(r'^(\d+)\s*(.*)$', adet_str)
                    if match:
                        adet = int(match.group(1))
                        birim_raw = match.group(2).strip()
                        if birim_raw:
                            # Birim dönüşümü
                            birim_map = {
                                'KUTU': 'Kutu',
                                'BİDON': 'Litre',
                                'BIDON': 'Litre',
                                'BOY': 'Adet',
                                'PAKET': 'Paket',
                                'ADET': 'Adet',
                                'CUVAL': 'Adet',
                                'ÇUVAL': 'Adet',
                                'KOVA': 'Litre',
                                'LITRE': 'Litre',
                                'LİTRE': 'Litre',
                                'METRE': 'Metre',
                                'KG': 'Kg',
                                'KILOGRAM': 'Kg'
                            }
                            birim = birim_map.get(birim_raw, 'Adet')
                    else:
                        try:
                            adet = int(float(adet_str))
                        except:
                            adet = 1
                
                # Yeni malzeme kodu oluştur
                new_code = self._generate_id("MAL")
                while new_code in existing_codes:
                    new_code = self._generate_id("MAL")
                
                # Kategori tahmin et (ürün adına göre)
                kategori = "Temizlik"  # Varsayılan
                urun_lower = urun_adi.lower()
                if any(k in urun_lower for k in ['kağıt', 'kalem', 'dosya', 'klasör', 'zımba', 'makas']):
                    kategori = "Kırtasiye"
                elif any(k in urun_lower for k in ['temizlik', 'çöp', 'deterjan', 'sabun', 'süpürge', 'paspas', 'bez', 'havlu']):
                    kategori = "Temizlik"
                elif any(k in urun_lower for k in ['kahve', 'çay', 'bardak', 'su', 'şeker']):
                    kategori = "Mutfak"
                elif any(k in urun_lower for k in ['boya', 'fırça', 'lambri', 'laminat', 'yapıştırıcı', 'silikon']):
                    kategori = "Teknik"
                elif any(k in urun_lower for k in ['toner', 'yazıcı', 'bilgisayar']):
                    kategori = "Ofis Ekipmanı"
                
                # Malzemeyi ekle
                ws.append([
                    new_code,
                    urun_adi,
                    kategori,
                    birim,
                    adet,
                    5,  # min_seviye
                    adet * 3 if adet > 0 else 100,  # max_seviye
                    "Depo",
                    "A-1",
                    "",  # barkod
                    0.0,  # birim_fiyat
                    now,
                    now
                ])
                
                existing_codes.add(new_code)
                existing_names.add(urun_adi.lower())
                imported_count += 1
            
            wb.save(self.file_path)
            wb.close()
            import_wb.close()
            
            return {
                "success": True,
                "imported": imported_count,
                "skipped": skipped_count,
                "skipped_items": skipped_items[:10],  # İlk 10 atlanan ürün
                "message": f"{imported_count} malzeme başarıyla içe aktarıldı"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e), "imported": 0, "skipped": 0}

    def get_monthly_stats(self) -> List[dict]:
        """Aylık istatistikler"""
        movements = self.get_all_movements()
        months = {}
        
        for m in movements:
            month = m.tarih[:7] if m.tarih and len(m.tarih) >= 7 else ""
            if not month:
                continue
            
            if month not in months:
                months[month] = {"giris": 0, "cikis": 0, "harcama": 0}
            
            if m.islem_tipi == "Giriş":
                months[month]["giris"] += m.miktar
            else:
                months[month]["cikis"] += m.miktar
        
        result = []
        for ay, data in sorted(months.items())[-12:]:
            result.append({
                "ay": ay,
                "giris": data["giris"],
                "cikis": data["cikis"],
                "harcama": data.get("harcama", 0)
            })
        return result

