"""
Script to add inventory data from the image to the Excel database
"""
from openpyxl import load_workbook
from datetime import datetime
import os

# Inventory data from the image
inventory_data = [
    {"ad": "Büyük Boy Çöp Torbası (Kutu)", "adet": 11, "birim": "Kutu"},
    {"ad": "Küçük Boy Çöp Torbası (Kutu)", "adet": 5, "birim": "Kutu"},
    {"ad": "Orta Boy Çöp Torbası", "adet": 4, "birim": "Adet"},
    {"ad": "Dispenser Uyumlu Tuvalet Kağıdı", "adet": 37, "birim": "Adet"},
    {"ad": "Karo Halı Zemin Yapıştırıcısı (Bidon)", "adet": 2, "birim": "Adet"},
    {"ad": "Temizlik Fırçası Ahşap Sap", "adet": 33, "birim": "Adet"},
    {"ad": "Alçıpan Levha", "adet": 8, "birim": "Adet"},
    {"ad": "Genel Amaçlı Temizlik Kimyasalı (Asperox)", "adet": 8, "birim": "Adet"},
    {"ad": "Yüzey Temizlik Havlusu", "adet": 3, "birim": "Adet"},
    {"ad": "Büyük Boy Boya Rulo Fırça Seti", "adet": 11, "birim": "Adet"},
    {"ad": "Cam Temizleyici (5 Litre)", "adet": 15, "birim": "Litre"},
    {"ad": "Çamaşır Suyu (18 Litre)", "adet": 3, "birim": "Litre"},
    {"ad": "Zemin Çekpas Başlığı", "adet": 21, "birim": "Adet"},
    {"ad": "Genel Amaçlı Silikon", "adet": 1, "birim": "Adet"},
    {"ad": "Kağıt El Peçetesi", "adet": 60, "birim": "Adet"},
    {"ad": "Döner Zemin Ovgu Makarası", "adet": 0, "birim": "Adet"},
    {"ad": "Süpürge – Faraş Seti", "adet": 45, "birim": "Adet"},
    {"ad": "Dispenser Uyumlu Kağıt El Havlusu", "adet": 18, "birim": "Adet"},
    {"ad": "Jumbo Büyük Çöp Torbası (Kutu)", "adet": 12, "birim": "Kutu"},
    {"ad": "Tek Kullanımlık Karton Bardak", "adet": 3, "birim": "Adet"},
    {"ad": "Endüstriyel Kireç Çözücü (30 Litre)", "adet": 2, "birim": "Litre"},
    {"ad": "Küçük Boy Boya Rulo Fırça Seti", "adet": 7, "birim": "Adet"},
    {"ad": "Dekoratif Duvar Kaplama (Lambiri)", "adet": 15, "birim": "Adet"},
    {"ad": "Laminat Zemin Kaplaması (Paket)", "adet": 4, "birim": "Paket"},
    {"ad": "Endüstriyel Makine Temizlik Sıvısı", "adet": 9, "birim": "Adet"},
    {"ad": "Plastik Temizlik Maşrapası", "adet": 25, "birim": "Adet"},
    {"ad": "Genel Amaçlı Temizlik Sıvısı – Mavi Su", "adet": 6, "birim": "Adet"},
    {"ad": "Büyük Boy Temizlik Bezi", "adet": 10, "birim": "Adet"},
    {"ad": "Metal Masaüstü Peçetelik", "adet": 2, "birim": "Adet"},
    {"ad": "Selülozik Tiner (3 Litre)", "adet": 5, "birim": "Litre"},
    {"ad": "Mop Tutucu Aparat Başlığı", "adet": 10, "birim": "Adet"},
    {"ad": "Küçük Boy Mop Bezi", "adet": 34, "birim": "Adet"},
    {"ad": "Esnek Musluk Bağlantı Hortumu", "adet": 2, "birim": "Adet"},
    {"ad": "Ofis Tipi Çöp Kovası (Siyah)", "adet": 1, "birim": "Adet"},
    {"ad": "Araç Yıkama Fırçası", "adet": 15, "birim": "Adet"},
    {"ad": "Temizlik Amaçlı Pacavra Bezi (Çuval)", "adet": 1, "birim": "Adet"},
    {"ad": "Pisuar Koku Giderici Tablet", "adet": 54, "birim": "Adet"},
    {"ad": "Endüstriyel Boya – RAL 7038 (18 L)", "adet": 1, "birim": "Adet"},
    {"ad": "Endüstriyel Boya – RAL 9002 Kırık Beyaz (18L)", "adet": 0.5, "birim": "Adet"},
    {"ad": "Saf / Deiyonize Su (1 Litre)", "adet": 8, "birim": "Litre"},
    {"ad": "İş Güvenliği Zemin İşaretleme Bandı", "adet": 9, "birim": "Adet"},
    {"ad": "Sıvı El Sabunu (20 Litre)", "adet": 5, "birim": "Litre"},
    {"ad": "Çok Amaçlı Silikon Yapıştırıcı", "adet": 4, "birim": "Adet"},
    {"ad": "Genel Amaçlı Temizlik Bezi", "adet": 47, "birim": "Adet"},
    {"ad": "Temizlik Süngeri", "adet": 3, "birim": "Adet"},
    {"ad": "Tel Bulaşık / Temizlik Teli", "adet": 10, "birim": "Adet"},
    {"ad": "Yer Temizlik Bezi (Vileda Tipi)", "adet": 3, "birim": "Adet"},
    {"ad": "Temizlik Kovası (Presli)", "adet": 9, "birim": "Adet"},
    {"ad": "WC Temizlik Kimyasalı (20 Litre)", "adet": 1, "birim": "Litre"},
    {"ad": "Yer Halısı (Kutu)", "adet": 1, "birim": "Kutu"},
    {"ad": "Tahta Çay Kaşığı", "adet": 7, "birim": "Kutu"},
]

def get_category(product_name):
    """Determine category based on product name"""
    name_lower = product_name.lower()
    
    if any(k in name_lower for k in ['temizlik', 'çöp', 'deterjan', 'sabun', 'süpürge', 'paspas', 'bez', 'mop', 'fırça', 'süngeri', 'teli', 'kova', 'pacavra', 'havlu', 'peçete', 'wc', 'kireç', 'cam temiz']):
        return "Temizlik"
    elif any(k in name_lower for k in ['boya', 'laminat', 'alçıpan', 'silikon', 'yapıştırıcı', 'tiner', 'zemin', 'kaplama', 'lambiri', 'levha', 'halı']):
        return "Teknik"
    elif any(k in name_lower for k in ['kağıt', 'kalem', 'dosya', 'bardak', 'kaşık']):
        return "Kırtasiye"
    elif any(k in name_lower for k in ['musluk', 'hortum', 'bant', 'kablo']):
        return "Teknik"
    elif any(k in name_lower for k in ['peçetelik', 'pisuar']):
        return "Ofis Ekipmanı"
    else:
        return "Temizlik"

def main():
    file_path = "inventory_data.xlsx"
    
    # Load existing workbook
    wb = load_workbook(file_path)
    ws = wb["Malzemeler"]
    
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    # Get existing material codes to avoid duplicates
    existing_codes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing_codes.add(row[0])
    
    # Counter for new codes
    added = 0
    skipped = 0
    
    for idx, item in enumerate(inventory_data, 1):
        # Check if item already exists by name
        item_exists = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] and row[1].lower() == item["ad"].lower():
                item_exists = True
                break
        
        if item_exists:
            skipped += 1
            continue
        
        # Generate unique code
        code = f"MK{idx:04d}"
        while code in existing_codes:
            idx += 100
            code = f"MK{idx:04d}"
        existing_codes.add(code)
        
        # Determine category
        kategori = get_category(item["ad"])
        
        # Calculate min/max levels based on current stock
        min_level = max(1, int(item["adet"] * 0.2))
        max_level = max(10, int(item["adet"] * 3))
        
        # Add new row: Kod, Ad, Kategori, Birim, Stok, Min, Max, Konum, Raf, Barkod, Fiyat, Son Güncelleme, Son Sayım
        ws.append([
            code,
            item["ad"],
            kategori,
            item["birim"],
            int(item["adet"]) if item["adet"] == int(item["adet"]) else item["adet"],
            min_level,
            max_level,
            "Ana Depo",  # Default location
            "",  # Raf
            "",  # Barkod
            0,   # Fiyat
            now,
            now
        ])
        added += 1
    
    # Save workbook
    wb.save(file_path)
    wb.close()
    
    print(f"✅ {added} malzeme başarıyla eklendi")
    print(f"⚠️ {skipped} malzeme zaten mevcut (atlandı)")

if __name__ == "__main__":
    main()
