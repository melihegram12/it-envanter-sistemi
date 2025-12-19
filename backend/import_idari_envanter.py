"""
İDARİ İŞLER ENVANTER.xlsx dosyasından verileri inventory_data.xlsx'e aktarır
"""
from openpyxl import load_workbook
from datetime import datetime
import re

def parse_quantity(adet_str):
    """Adet stringini sayıya çevir: '10 ADET', '5 KUTU' vb."""
    if adet_str is None:
        return 0, "Adet"
    
    adet_str = str(adet_str).strip()
    
    # Sayıyı bul
    match = re.search(r'(\d+)', adet_str)
    miktar = int(match.group(1)) if match else 0
    
    # Birimi bul
    birim = "Adet"
    adet_lower = adet_str.lower()
    if "paket" in adet_lower:
        birim = "Paket"
    elif "kutu" in adet_lower:
        birim = "Kutu"
    elif "litre" in adet_lower or "lt" in adet_lower or "bidon" in adet_lower:
        birim = "Litre"
    elif "kg" in adet_lower or "kilo" in adet_lower:
        birim = "Kg"
    elif "koli" in adet_lower:
        birim = "Koli"
    elif "takım" in adet_lower or "set" in adet_lower:
        birim = "Adet"
    elif "top" in adet_lower or "rulo" in adet_lower:
        birim = "Adet"
    
    return miktar, birim

def determine_category(urun_adi):
    """Ürün adından kategori belirle"""
    urun_lower = urun_adi.lower() if urun_adi else ""
    
    if any(x in urun_lower for x in ["boya", "fırça", "rulo", "rulosu", "macun", "spatula", "astar"]):
        return "Teknik"
    elif any(x in urun_lower for x in ["temiz", "deterjan", "sabun", "çamaşır", "bulaşık", "hijyen", "kova", "paspas", "süpürge"]):
        return "Temizlik"
    elif any(x in urun_lower for x in ["çay", "kahve", "bardak", "şeker", "tabak", "kaşık", "peçete", "yemek"]):
        return "Mutfak"
    elif any(x in urun_lower for x in ["kalem", "kağıt", "dosya", "zımba", "makas", "silgi", "cetvel", "defter"]):
        return "Kırtasiye"
    elif any(x in urun_lower for x in ["halı", "zemin", "halıfleks", "karo"]):
        return "Teknik"
    else:
        return "Diğer"

def main():
    source_path = r'C:\Users\ENGINME1\Desktop\İDARİ İŞLER ENVANTER.xlsx'
    target_path = 'inventory_data.xlsx'
    
    # Kaynak Excel'i oku
    print("Kaynak dosya okunuyor...")
    source_wb = load_workbook(source_path)
    source_ws = source_wb.active
    
    # Hedef Excel'i oku
    print("Hedef dosya okunuyor...")
    target_wb = load_workbook(target_path)
    target_ws = target_wb["Malzemeler"]
    
    # Mevcut malzeme kodlarını al (tekrar eklemeyi önlemek için)
    existing_codes = set()
    for row in target_ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing_codes.add(row[0])
    
    print(f"Mevcut malzeme sayısı: {len(existing_codes)}")
    
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    imported_count = 0
    skipped_count = 0
    
    # Veri satırlarını oku (3. satırdan başla)
    for i in range(3, source_ws.max_row + 1):
        sayi = source_ws.cell(i, 2).value  # B sütunu - Sayı
        urun_adi = source_ws.cell(i, 3).value  # C sütunu - Ürün Adı
        adet = source_ws.cell(i, 4).value  # D sütunu - Adet
        durum = source_ws.cell(i, 5).value  # E sütunu - Durum (P/V gibi)
        
        if not urun_adi:
            continue
        
        # Malzeme kodu oluştur
        kod = f"ENV{str(sayi).zfill(3)}"
        
        # Zaten varsa atla
        if kod in existing_codes:
            skipped_count += 1
            continue
        
        # Adet ve birimi parse et
        miktar, birim = parse_quantity(adet)
        
        # Kategoriyi belirle
        kategori = determine_category(urun_adi)
        
        # Durumu belirle
        stok_durum = "Normal" if durum == "P" else "Kritik" if durum == "V" else "Normal"
        
        # Minimum seviyeyi tahmin et
        min_seviye = max(1, miktar // 5)
        max_seviye = max(miktar * 2, 50)
        
        # Hedef Excel'e ekle
        target_ws.append([
            kod,                    # Kod
            str(urun_adi).strip(),  # Ad
            kategori,               # Kategori
            birim,                  # Birim
            miktar,                 # Mevcut Stok
            min_seviye,             # Min Seviye
            max_seviye,             # Max Seviye
            "Depo",                 # Konum
            "A-1",                  # Raf
            "",                     # Barkod
            0.0,                    # Birim Fiyat
            now,                    # Son Güncelleme
            now                     # Son Sayım
        ])
        
        existing_codes.add(kod)
        imported_count += 1
        print(f"  + {kod}: {urun_adi} ({miktar} {birim}) - {kategori}")
    
    # Kaydet
    target_wb.save(target_path)
    print()
    print(f"=== SONUÇ ===")
    print(f"Aktarılan: {imported_count} malzeme")
    print(f"Atlanan (zaten var): {skipped_count} malzeme")
    print(f"Toplam malzeme: {len(existing_codes)}")
    print("İşlem tamamlandı!")

if __name__ == "__main__":
    main()
